"""Import 24-25 season data from Excel into matches.json for the archive site."""

import csv
import json
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path
from match_results import RESULTS

SPECIAL_ENTRIES = {
    "Rating",
    "Referee Rating",
    "Postecoglou - Overall Rating",
    "Postecoglou - Starting 11 Selection",
    "Postecoglou - Tactics",
    "Postecoglou - Use of Subs",
}


def is_special(name: str) -> bool:
    if name == "Referee Rating":
        return True
    if name.startswith("Postecoglou"):
        return True
    if name.endswith(" Rating"):
        return True
    return False


def excel_date(val) -> str:
    if isinstance(val, (int, float)):
        d = datetime(1899, 12, 30) + timedelta(days=int(val))
        return d.strftime("%Y-%m-%d")
    return str(val)[:10]


def build_results_lookup() -> dict[tuple[str, str], tuple]:
    """Build a lookup from (date, opponent) -> (spurs_score, opp_score, is_home, venue).

    Scores are always Tottenham first, opponent second.
    is_home can be True, False, or None (neutral venue).
    """
    lookup = {}
    for date, opp, spurs_score, opp_score, is_home, venue in RESULTS:
        lookup[(date, opp)] = (spurs_score, opp_score, is_home, venue)
    return lookup


def main():
    wb = openpyxl.load_workbook("24-25_seasondata.xlsx")
    ws = wb["Sheet1"]

    results_lookup = build_results_lookup()

    # Group rows by match (date + opponent)
    raw_matches: dict[tuple, list] = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date_val, comp, rnd, opp, name, rating, stdev = row
        if date_val is None or name is None:
            continue
        date_str = excel_date(date_val)
        key = (date_str, opp)
        if key not in raw_matches:
            raw_matches[key] = {"comp": comp, "rnd": rnd, "rows": []}
        raw_matches[key]["rows"].append((name, rating, stdev))

    matches_index = []
    all_match_details = []

    for i, ((date_str, opponent), info) in enumerate(sorted(raw_matches.items()), start=1):
        match_id = i
        comp = info["comp"]
        rnd = info["rnd"]
        rows = info["rows"]

        players = []
        team_rating = None
        opp_rating = None
        referee_rating = None
        coach_overall = None
        coach_selection = None
        coach_tactics = None
        coach_subs = None

        for name, rating, stdev in rows:
            r = float(rating) if rating else 0.0
            s = float(stdev) if stdev else 0.0

            if name == "Tottenham Rating":
                team_rating = {"mean": r, "std_dev": s}
            elif name == "Referee Rating":
                referee_rating = {"mean": r, "std_dev": s}
            elif name == "Postecoglou - Overall Rating":
                coach_overall = {"mean": r, "std_dev": s}
            elif name == "Postecoglou - Starting 11 Selection":
                coach_selection = {"mean": r, "std_dev": s}
            elif name == "Postecoglou - Tactics":
                coach_tactics = {"mean": r, "std_dev": s}
            elif name == "Postecoglou - Use of Subs":
                coach_subs = {"mean": r, "std_dev": s}
            elif name.endswith(" Rating"):
                opp_rating = {"mean": r, "std_dev": s}
            else:
                players.append({
                    "name": name,
                    "position": "",
                    "image_path": None,
                    "rating": {"mean": r, "std_dev": s},
                    "is_starter": True,
                    "goals": 0,
                    "assists": 0,
                    "own_goals": 0,
                    "is_motm": False,
                })

        # Compute average player rating
        if players:
            avg = sum(p["rating"]["mean"] for p in players) / len(players)
        else:
            avg = 0.0

        # MOTM = highest rated player
        if players:
            top = max(players, key=lambda p: p["rating"]["mean"])
            top["is_motm"] = True
            motm_name = top["name"]
        else:
            motm_name = ""

        defaults = {"mean": 0.0, "std_dev": 0.0}

        # Look up actual match result
        result_data = results_lookup.get((date_str, opponent))
        if result_data:
            spurs_score, opp_score, is_home, venue = result_data
        else:
            print(f"  WARNING: No result found for {date_str} vs {opponent}")
            spurs_score, opp_score, is_home, venue = 0, 0, True, ""

        # home_away: "Home", "Away", or "Neutral"
        if is_home is None:
            home_away = "Neutral"
        elif is_home:
            home_away = "Home"
        else:
            home_away = "Away"

        detail = {
            "match_id": match_id,
            "metadata": {
                "opponent": opponent,
                "competition": comp,
                "matchday": rnd,
                "date": date_str,
                "venue": venue,
                "spurs_score": spurs_score,
                "opponent_score": opp_score,
                "home_away": home_away,
            },
            "formation": "",
            "team_rating": team_rating or defaults,
            "opponent_rating": opp_rating or defaults,
            "referee_rating": referee_rating or defaults,
            "coach_ratings": {
                "name": "Postecoglou",
                "starting_eleven": coach_selection or defaults,
                "on_field_tactics": coach_tactics or defaults,
                "substitutions": coach_subs or defaults,
            },
            "overall_rating": round(avg, 2),
            "starting_players": [p for p in players],
            "substitute_players": [],
            "motm_winners": [motm_name] if motm_name else [],
            "total_responses": 0,
        }

        all_match_details.append(detail)

        matches_index.append({
            "match_id": match_id,
            "opponent": opponent,
            "competition": comp,
            "matchday": rnd,
            "date": date_str,
            "venue": venue,
            "spurs_score": spurs_score,
            "opponent_score": opp_score,
            "home_away": home_away,
        })

    # Sort most recent first
    matches_index.sort(key=lambda m: m["date"], reverse=True)

    # Write index
    out = Path("archive-app/public/data")
    out.mkdir(parents=True, exist_ok=True)

    with open(out / "matches.json", "w") as f:
        json.dump(matches_index, f, indent=2)

    # Write individual match details
    matches_dir = out / "matches"
    matches_dir.mkdir(exist_ok=True)
    for detail in all_match_details:
        mid = detail["match_id"]
        mdir = matches_dir / str(mid)
        mdir.mkdir(exist_ok=True)
        with open(mdir / "results.json", "w") as f:
            json.dump(detail, f, indent=2)

    print(f"Wrote {len(matches_index)} matches to {out / 'matches.json'}")
    print(f"Wrote {len(all_match_details)} match detail files")

    # Write CSV with all player ratings flattened
    csv_path = out / "all_ratings.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "date", "opponent", "competition", "matchday", "venue",
            "home_away", "spurs_score", "opponent_score", "result",
            "player", "rating", "std_dev", "is_motm",
            "team_rating", "team_rating_std",
            "opponent_rating", "opponent_rating_std",
            "referee_rating", "referee_rating_std",
            "overall_rating",
        ])
        for detail in sorted(all_match_details, key=lambda d: d["metadata"]["date"], reverse=True):
            m = detail["metadata"]
            ss, os_ = m["spurs_score"], m["opponent_score"]
            result = "W" if ss > os_ else ("L" if ss < os_ else "D")
            tr = detail["team_rating"]
            opr = detail["opponent_rating"]
            rr = detail["referee_rating"]
            for p in detail["starting_players"] + detail["substitute_players"]:
                writer.writerow([
                    m["date"], m["opponent"], m["competition"], m["matchday"],
                    m["venue"], m["home_away"], ss, os_, result,
                    p["name"], round(p["rating"]["mean"], 2), round(p["rating"]["std_dev"], 2),
                    p["is_motm"],
                    round(tr["mean"], 2), round(tr["std_dev"], 2),
                    round(opr["mean"], 2), round(opr["std_dev"], 2),
                    round(rr["mean"], 2), round(rr["std_dev"], 2),
                    detail["overall_rating"],
                ])
    print(f"Wrote CSV to {csv_path}")


if __name__ == "__main__":
    main()
